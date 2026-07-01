import argparse
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_API_URL = "http://127.0.0.1:8000"
DEFAULT_DAYS = 7
DEFAULT_OUTPUT_DIR = "reports"
REQUEST_TIMEOUT_SECONDS = 20
COLORS = ("2f6f9f", "c9564a", "6f8f3f", "d39b32", "7259a4", "3f8f88", "8a5c3b", "69717c")
STATION_DISPLAY_NAMES = {
    "ART_ENDFORM_1859": "Tesla 1",
    "ART_ENDFORM_1861": "Tesla 2",
    "ART_ENDFORM_1862": "Tesla 3",
}


class ReportError(RuntimeError):
    pass


@dataclass(frozen=True)
class ReportParams:
    api_url: str
    start_at: str
    end_at: str
    source_station: str | None = None
    part_numbers: list[str] | None = None
    output_dir: str = DEFAULT_OUTPUT_DIR


def parse_datetime(value, name):
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise ReportError(f"{name} must use ISO datetime format") from exc


def default_period(days=DEFAULT_DAYS):
    end_at = datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)
    start_at = (end_at - timedelta(days=days)).replace(hour=0, minute=0, second=0, microsecond=0)
    return start_at, end_at


def build_params(args):
    if args.start_at or args.end_at:
        if not args.start_at or not args.end_at:
            raise ReportError("--start-at and --end-at must be used together")
        start_at = parse_datetime(args.start_at, "--start-at")
        end_at = parse_datetime(args.end_at, "--end-at")
    else:
        start_at, end_at = default_period(args.days)

    if end_at < start_at:
        raise ReportError("--end-at must be greater than or equal to --start-at")

    return ReportParams(
        api_url=args.api_url.rstrip("/"),
        start_at=start_at.isoformat(sep=" ", timespec="seconds"),
        end_at=end_at.isoformat(sep=" ", timespec="seconds"),
        source_station=args.source_station,
        part_numbers=args.part_numbers or None,
        output_dir=args.output_dir,
    )


def _request_json(session, url, params=None):
    try:
        response = session.get(url, params=params, timeout=REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as exc:
        raise ReportError(f"API request failed: {url} ({exc})") from exc
    except ValueError as exc:
        raise ReportError(f"API returned invalid JSON: {url}") from exc


def fetch_report_data(report_params, session=None):
    session = session or requests.Session()
    health = _request_json(session, f"{report_params.api_url}/health")
    if health.get("api") != "ok" or health.get("database") != "ok":
        detail = health.get("detail") or health
        raise ReportError(f"API health check failed: {detail}")

    params = {
        "start_at": report_params.start_at,
        "end_at": report_params.end_at,
    }
    if report_params.source_station:
        params["source_station"] = report_params.source_station
    if report_params.part_numbers:
        params["part_numbers"] = report_params.part_numbers

    return _request_json(
        session,
        f"{report_params.api_url}/api/v1/reject-summary",
        params=params,
    )


def _normalize_number(value, default=0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_int(value):
    return int(_normalize_number(value, 0))


def _date_label(value):
    return str(value or "")[:10]


def _station_name(value):
    text = str(value or "").strip()
    if not text:
        return "Sin estacion"

    match = re.search(r"(?:\s*-\s*|_)(LEFT|RIGHT)$", text, re.IGNORECASE)
    raw_base = text[: match.start()].strip(" _-") if match else text
    display_base = STATION_DISPLAY_NAMES.get(raw_base)
    if not display_base:
        return text
    if not match:
        return display_base

    side = match.group(1).upper()
    return f"{display_base} - {'Left' if side == 'LEFT' else 'Right'}"


def _defect_name(value):
    text = str(value or "").strip().upper()
    return text or "UNCLASSIFIED"


def _report_defect_names(data):
    names = set()
    for collection in ("condition_totals", "condition_periods", "top3_history"):
        for row in data.get(collection) or []:
            name = _defect_name(row.get("class_name"))
            if name != "OK":
                names.add(name)
    return sorted(names)


def _defect_color_map(data):
    return {name: COLORS[idx % len(COLORS)] for idx, name in enumerate(_report_defect_names(data))}


def _defect_color(colors_by_defect, name):
    return colors_by_defect.get(_defect_name(name), COLORS[0])


def _nice_axis_max(values):
    max_value = max([_normalize_int(value) for value in values or []] or [0])
    if max_value <= 0:
        return 1
    magnitude = 10 ** (len(str(max_value)) - 1)
    normalized = max_value / magnitude
    if normalized <= 1:
        nice_normalized = 1
    elif normalized <= 2:
        nice_normalized = 2
    elif normalized <= 5:
        nice_normalized = 5
    else:
        nice_normalized = 10
    return nice_normalized * magnitude


def _safe_sheet_title(title):
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", str(title or "Sheet")).strip()
    return (cleaned or "Sheet")[:31]


def _safe_table_name(name):
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", str(name or "Table"))
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"tbl_{cleaned}"
    return cleaned[:200]


def _style_header(sheet, row_idx, min_col=1, max_col=None):
    max_col = max_col or sheet.max_column
    header_fill = PatternFill(fill_type="solid", fgColor="404040")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(min_col, max_col + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _fit_columns(sheet, max_width=42):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(width + 2, 12), max_width)


def _add_table(sheet, table_name, header_row, min_col, max_row, max_col):
    ref = f"{sheet.cell(row=header_row, column=min_col).coordinate}:{sheet.cell(row=max_row, column=max_col).coordinate}"
    table = Table(displayName=_safe_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _append_table(sheet, table_name, headers, rows, start_row=None, start_col=1):
    start_row = start_row or sheet.max_row + 1
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        start_row = 1
    rows = list(rows or [])
    if not rows:
        rows = [["Sin datos", *[""] * (len(headers) - 1)]]

    for offset, value in enumerate(headers):
        sheet.cell(row=start_row, column=start_col + offset, value=value)
    _style_header(sheet, start_row, start_col, start_col + len(headers) - 1)

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_offset, value in enumerate(row):
            sheet.cell(row=row_idx, column=start_col + col_offset, value=value)

    max_row = start_row + len(rows)
    max_col = start_col + len(headers) - 1
    _add_table(sheet, table_name, start_row, start_col, max_row, max_col)
    return start_row, max_row, max_col


def _group_by(rows, key):
    grouped = {}
    for row in rows or []:
        group_key = row.get(key) or ""
        grouped.setdefault(group_key, []).append(row)
    return grouped


def _station_list(data):
    names = set()
    for collection in ("stations", "daily", "condition_periods", "condition_totals", "top3_history"):
        for row in data.get(collection) or []:
            names.add(row.get("source_station") or "")
    return sorted(names, key=_station_name)


def _condition_station_list(data):
    names = {row.get("source_station") or "" for row in data.get("condition_periods") or []}
    return sorted(names, key=_station_name)


def _combined_as_station_data(data):
    combined = (data or {}).get("combined") or {}

    def map_rows(rows):
        return [
            {
                **row,
                "source_station": row.get("station_pair") or row.get("source_station") or "",
            }
            for row in rows or []
        ]

    return {
        "stations": map_rows(combined.get("stations")),
        "daily": map_rows(combined.get("daily")),
        "condition_periods": map_rows(combined.get("condition_periods")),
        "condition_totals": map_rows(combined.get("condition_totals")),
        "top3_history": map_rows(combined.get("top3_history")),
    }


def _condition_classes(rows):
    totals = {}
    for row in rows or []:
        class_name = _defect_name(row.get("class_name"))
        nok_pieces = _normalize_int(row.get("nok_pieces"))
        if class_name == "OK" or nok_pieces <= 0:
            continue
        totals[class_name] = totals.get(class_name, 0) + nok_pieces
    return sorted(totals)


def _condition_daily_rows(rows, classes):
    by_date = {}
    for row in rows or []:
        class_name = _defect_name(row.get("class_name"))
        nok_pieces = _normalize_int(row.get("nok_pieces"))
        if class_name == "OK" or nok_pieces <= 0:
            continue
        day = _date_label(row.get("reject_date"))
        if not day:
            continue
        by_date.setdefault(day, {"reject_date": day, "total_nok": 0})
        by_date[day][class_name] = by_date[day].get(class_name, 0) + nok_pieces
        by_date[day]["total_nok"] += nok_pieces

    output = []
    for day in sorted(by_date):
        item = by_date[day]
        output.append([day, *[_normalize_int(item.get(name)) for name in classes], _normalize_int(item["total_nok"])])
    return output


def _top_history_rows(rows, classes):
    by_date = {}
    for row in rows or []:
        day = _date_label(row.get("reject_date"))
        class_name = _defect_name(row.get("class_name"))
        if not day:
            continue
        by_date.setdefault(day, {"reject_date": day})
        by_date[day][class_name] = _normalize_int(row.get("nok_pieces"))

    output = []
    for day in sorted(by_date):
        item = by_date[day]
        output.append([day, *[_normalize_int(item[name]) if name in item else "" for name in classes]])
    return output


def _append_condition_daily_table(sheet, table_name, headers, daily_rows, start_row):
    header_row, daily_max_row, daily_max_col = _append_table(
        sheet,
        table_name,
        headers,
        daily_rows,
        start_row=start_row,
    )
    if daily_max_col <= 1:
        return header_row, daily_max_row, daily_max_col, None

    total_row = daily_max_row + 1
    sheet.cell(row=total_row, column=1, value="Total")
    sheet.cell(row=total_row, column=1).font = Font(bold=True)
    for col_idx in range(2, daily_max_col + 1):
        col_letter = get_column_letter(col_idx)
        cell = sheet.cell(row=total_row, column=col_idx)
        cell.value = f"=SUM({col_letter}{header_row + 1}:{col_letter}{daily_max_row})"
        cell.font = Font(bold=True)

    return header_row, daily_max_row, daily_max_col, total_row


def _condition_daily_meta(sheet, station, label, classes, daily_rows, header_row, daily_max_row, daily_max_col, total_row):
    totals = {}
    values = {}
    columns = {}
    for class_offset, class_name in enumerate(classes, start=1):
        col_idx = 1 + class_offset
        class_values = [
            _normalize_int(row[class_offset]) if class_offset < len(row) else 0
            for row in daily_rows
        ]
        columns[class_name] = col_idx
        values[class_name] = class_values
        totals[class_name] = sum(class_values)

    return {
        "sheet": sheet,
        "station": station,
        "label": label,
        "classes": list(classes),
        "columns": columns,
        "values": values,
        "totals": totals,
        "header_row": header_row,
        "data_start_row": header_row + 1,
        "data_end_row": daily_max_row,
        "daily_max_col": daily_max_col,
        "total_row": total_row,
    }


def _add_condition_pie_chart(sheet, label, classes, header_row, total_row, colors_by_defect, anchor):
    if not classes or not total_row:
        return

    class_end_col = 1 + len(classes)
    chart = PieChart()
    chart.title = f"{label} - Rechazos por clase"
    chart.height = 8
    chart.width = 12
    chart.add_data(
        Reference(sheet, min_col=2, max_col=class_end_col, min_row=total_row, max_row=total_row),
        from_rows=True,
    )
    chart.set_categories(Reference(sheet, min_col=2, max_col=class_end_col, min_row=header_row, max_row=header_row))
    chart.series[0].data_points = [
        DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=_defect_color(colors_by_defect, class_name)))
        for idx, class_name in enumerate(classes)
    ]
    sheet.add_chart(chart, anchor)


def _top_condition_classes(condition_meta, limit=3):
    totals = condition_meta.get("totals") or {}
    return [
        class_name
        for class_name, _total in sorted(totals.items(), key=lambda item: (-item[1], item[0]))[:limit]
    ]


def _top_condition_axis_max(condition_meta, classes):
    values = []
    for class_name in classes:
        values.extend(condition_meta.get("values", {}).get(class_name, []))
    return _nice_axis_max(values)


def _add_top3_condition_chart(sheet, condition_meta, classes, colors_by_defect, anchor):
    if not classes:
        return

    condition_sheet = condition_meta["sheet"]
    chart = BarChart()
    chart.title = f"{condition_meta['label']} - Top 3 historico"
    chart.y_axis.title = "NOK"
    chart.x_axis.title = "Fecha"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = _top_condition_axis_max(condition_meta, classes)
    chart.y_axis.number_format = "0"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height = 8
    chart.width = 16

    for class_name in classes:
        col_idx = condition_meta["columns"][class_name]
        chart.add_data(
            Reference(
                condition_sheet,
                min_col=col_idx,
                max_col=col_idx,
                min_row=condition_meta["header_row"],
                max_row=condition_meta["data_end_row"],
            ),
            titles_from_data=True,
        )
        chart.series[-1].tx = SeriesLabel(v=class_name)
        chart.series[-1].graphicalProperties.solidFill = _defect_color(colors_by_defect, class_name)

    chart.set_categories(
        Reference(
            condition_sheet,
            min_col=1,
            min_row=condition_meta["data_start_row"],
            max_row=condition_meta["data_end_row"],
        )
    )
    sheet.add_chart(chart, anchor)


def _format_percentage_columns(sheet, columns, start_row=2):
    for col_idx in columns:
        for row_idx in range(start_row, sheet.max_row + 1):
            sheet.cell(row=row_idx, column=col_idx).number_format = "0.0%"


DAILY_METRICS = ("OK", "NOK", "Total", "% OK", "% NOK")
DAILY_STANDARD_BLOCKS = ("Left", "Right", "Combinado")
DAILY_FILTER_ROWS = 4


def _filter_value(value):
    text = str(value or "").strip()
    return text or "Todos"


def _filter_list_value(values):
    cleaned = [str(value).strip() for value in values or [] if str(value or "").strip()]
    return ", ".join(cleaned) if cleaned else "Todos"


def _write_filter_band(sheet, report_params):
    title_fill = PatternFill(fill_type="solid", fgColor="E8EEF5")
    label_font = Font(bold=True)
    rows = [
        ("Filtros aplicados", ""),
        (f"Inicio: {_filter_value(report_params.start_at)}", f"Fin: {_filter_value(report_params.end_at)}"),
        (f"Estacion: {_filter_value(report_params.source_station)}", f"Part Number: {_filter_list_value(report_params.part_numbers)}"),
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet.cell(row=1, column=1, value=rows[0][0])
    sheet.cell(row=1, column=1).fill = title_fill
    sheet.cell(row=1, column=1).font = label_font
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, (left, right) in enumerate(rows[1:], start=2):
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        sheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
        sheet.cell(row=row_idx, column=1, value=left)
        sheet.cell(row=row_idx, column=3, value=right)
        sheet.cell(row=row_idx, column=1).font = label_font
        sheet.cell(row=row_idx, column=3).font = label_font


def _split_station_side(value):
    text = _station_name(value).strip()
    match = re.search(r"(?:\s*-\s*|_)(LEFT|RIGHT)$", text, re.IGNORECASE)
    if not match:
        return text, ""

    base = text[: match.start()].strip(" _-")
    side = match.group(1).upper()
    return base or text, "Left" if side == "LEFT" else "Right"


def _daily_block_value(row):
    if not row:
        return ["", "", "", "", ""]
    return [
        _normalize_int(row.get("ok_pieces")),
        _normalize_int(row.get("nok_pieces")),
        _normalize_int(row.get("total_pieces")),
        _normalize_number(row.get("pct_ok")),
        _normalize_number(row.get("pct_nok")),
    ]


def _daily_layout(data):
    daily_rows = data.get("daily") or []
    combined_rows = ((data.get("combined") or {}).get("daily")) or []
    groups = {}
    by_block_date = {}
    dates = set()

    for row in daily_rows:
        day = _date_label(row.get("reject_date"))
        if not day:
            continue
        base, side = _split_station_side(row.get("source_station") or "")
        if not base:
            continue
        block = side or _station_name(row.get("source_station"))
        dates.add(day)
        groups.setdefault(base, {"blocks": set(), "has_standard_side": False})
        groups[base]["blocks"].add(block)
        groups[base]["has_standard_side"] = groups[base]["has_standard_side"] or block in {"Left", "Right"}
        by_block_date[(base, block, day)] = row

    for row in combined_rows:
        day = _date_label(row.get("reject_date"))
        if not day:
            continue
        base = row.get("station_pair") or _split_station_side(row.get("source_station") or "")[0]
        base = _station_name(base).strip()
        if not base:
            continue
        dates.add(day)
        groups.setdefault(base, {"blocks": set(), "has_standard_side": False})
        groups[base]["blocks"].add("Combinado")
        by_block_date[(base, "Combinado", day)] = row

    ordered_groups = []
    for base in sorted(groups, key=_station_name):
        info = groups[base]
        if info["has_standard_side"] or "Combinado" in info["blocks"]:
            blocks = [block for block in DAILY_STANDARD_BLOCKS if block in info["blocks"] or block != "Combinado"]
            if "Combinado" not in blocks:
                blocks.append("Combinado")
        else:
            blocks = sorted(info["blocks"], key=_station_name)
        ordered_groups.append((base, blocks))

    return sorted(dates), ordered_groups, by_block_date


def _apply_daily_border(sheet, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="000000")
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            left = thin if col_idx == min_col else cell.border.left
            right = thin if col_idx == max_col else cell.border.right
            top = thin if row_idx == min_row else cell.border.top
            bottom = thin if row_idx == max_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _write_daily_sheet(workbook, data, report_params):
    sheet = workbook.active
    sheet.title = "Por dia"
    _write_filter_band(sheet, report_params)
    dates, groups, by_block_date = _daily_layout(data)
    header_fill = PatternFill(fill_type="solid", fgColor="404040")
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    percent_cols = []
    nok_percent_cols = []
    series_cols = []
    group_row = DAILY_FILTER_ROWS + 1
    block_row = group_row + 1
    metric_row = group_row + 2
    data_start_row = group_row + 3

    sheet.cell(row=metric_row, column=1, value="Date")
    sheet.cell(row=metric_row, column=1).fill = header_fill
    sheet.cell(row=metric_row, column=1).font = header_font
    sheet.cell(row=metric_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    _apply_daily_border(sheet, group_row, metric_row, 1, 1)

    col_idx = 2
    for base, blocks in groups:
        base_start_col = col_idx
        for block in blocks:
            block_start_col = col_idx
            block_end_col = col_idx + len(DAILY_METRICS) - 1
            sheet.merge_cells(start_row=block_row, start_column=block_start_col, end_row=block_row, end_column=block_end_col)
            sheet.cell(row=block_row, column=block_start_col, value=block)
            sheet.cell(row=block_row, column=block_start_col).alignment = Alignment(horizontal="center", vertical="center")

            for metric_offset, metric in enumerate(DAILY_METRICS):
                metric_col = block_start_col + metric_offset
                cell = sheet.cell(row=metric_row, column=metric_col, value=metric)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if metric.startswith("%"):
                    percent_cols.append(metric_col)
                if metric == "% NOK":
                    nok_percent_cols.append(metric_col)
                    series_cols.append((base, block, metric_col))

            _apply_daily_border(sheet, block_row, metric_row, block_start_col, block_end_col)
            col_idx = block_end_col + 1

        base_end_col = col_idx - 1
        sheet.merge_cells(start_row=group_row, start_column=base_start_col, end_row=group_row, end_column=base_end_col)
        sheet.cell(row=group_row, column=base_start_col, value=base)
        sheet.cell(row=group_row, column=base_start_col).alignment = Alignment(horizontal="center", vertical="center")
        _apply_daily_border(sheet, group_row, group_row, base_start_col, base_end_col)

    for row_idx, day in enumerate(dates, start=data_start_row):
        sheet.cell(row=row_idx, column=1, value=day)
        sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        col_idx = 2
        for base, blocks in groups:
            for block in blocks:
                row = by_block_date.get((base, block, day))
                for offset, value in enumerate(_daily_block_value(row)):
                    cell = sheet.cell(row=row_idx, column=col_idx + offset, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += len(DAILY_METRICS)

    data_end_row = data_start_row + len(dates) - 1
    max_col = max(col_idx - 1, 1)
    sheet.freeze_panes = f"A{data_start_row}"

    if dates:
        for col in percent_cols:
            for row_idx in range(data_start_row, data_end_row + 1):
                sheet.cell(row=row_idx, column=col).number_format = "0.00%"
        for col in nok_percent_cols:
            avg_cell = sheet.cell(row=data_end_row + 1, column=col)
            letter = get_column_letter(col)
            avg_cell.value = f"=AVERAGE({letter}{data_start_row}:{letter}{data_end_row})"
            avg_cell.number_format = "0.00%"
            avg_cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.conditional_formatting.add(
            f"A{data_start_row}:{get_column_letter(max_col)}{data_end_row}",
            FormulaRule(formula=["ISEVEN(ROW())"], fill=stripe_fill),
        )
        for col in nok_percent_cols:
            letter = get_column_letter(col)
            sheet.conditional_formatting.add(
                f"{letter}{data_start_row}:{letter}{data_end_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )
        _apply_daily_border(sheet, data_start_row, data_end_row, 1, max_col)

    if not groups:
        sheet.cell(row=data_start_row, column=1, value="Sin datos")
        _fit_columns(sheet)
        return sheet

    if dates and series_cols:
        chart = LineChart()
        chart.title = "Tasa de rechazo (% NOK) por dia"
        chart.y_axis.title = "% NOK"
        chart.x_axis.title = "Fecha"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.height = 7.5
        chart.width = 15
        categories = Reference(sheet, min_col=1, min_row=data_start_row, max_row=data_end_row)
        for base, block, col_idx in series_cols:
            values = Reference(sheet, min_col=col_idx, max_col=col_idx, min_row=metric_row, max_row=data_end_row)
            chart.add_data(values, titles_from_data=True)
            chart.series[-1].tx = SeriesLabel(v=f"{base} - {block}")
        chart.set_categories(categories)
        sheet.add_chart(chart, f"A{data_end_row + 3}")

    _fit_columns(sheet)
    return sheet


def _append_combined_daily_section(sheet, data):
    combined_data = _combined_as_station_data(data)
    if not any(combined_data.values()):
        return

    row_cursor = sheet.max_row + 3
    sheet.cell(row=row_cursor, column=1, value="Combinado LEFT+RIGHT - Por dia")
    sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
    row_cursor += 1

    stations = _station_list(combined_data)
    daily_rows = combined_data.get("daily") or []
    dates = sorted({_date_label(row.get("reject_date")) for row in daily_rows if _date_label(row.get("reject_date"))})
    by_station_date = {
        (row.get("source_station") or "", _date_label(row.get("reject_date"))): row
        for row in daily_rows
    }

    headers = ["Fecha"]
    for station in stations:
        label = _station_name(station)
        headers.extend([f"{label} OK", f"{label} NOK", f"{label} Total", f"{label} % OK", f"{label} % NOK"])

    rows = []
    for day in dates:
        output = [day]
        for station in stations:
            row = by_station_date.get((station, day))
            if row:
                output.extend(
                    [
                        _normalize_int(row.get("ok_pieces")),
                        _normalize_int(row.get("nok_pieces")),
                        _normalize_int(row.get("total_pieces")),
                        _normalize_number(row.get("pct_ok")),
                        _normalize_number(row.get("pct_nok")),
                    ]
                )
            else:
                output.extend([0, 0, 0, "", ""])
        rows.append(output)

    if not headers[1:]:
        headers = ["Fecha", "Sin datos"]
    header_row, max_row, _ = _append_table(sheet, "tblPorDiaCombinado", headers, rows, start_row=row_cursor)
    percent_cols = [5 + (idx * 5) for idx in range(len(stations))]
    percent_cols.extend([6 + (idx * 5) for idx in range(len(stations))])
    _format_percentage_columns(sheet, percent_cols, start_row=header_row + 1)

    if rows and stations:
        chart = LineChart()
        chart.title = "Combinado LEFT+RIGHT - Tasa de rechazo (% NOK) por dia"
        chart.y_axis.title = "% NOK"
        chart.x_axis.title = "Fecha"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.height = 9
        chart.width = 18
        categories = Reference(sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
        for idx, _station in enumerate(stations):
            col_idx = 6 + (idx * 5)
            values = Reference(sheet, min_col=col_idx, max_col=col_idx, min_row=header_row, max_row=max_row)
            chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, f"{get_column_letter(len(headers) + 2)}{header_row}")


def _write_conditions_sheet(workbook, data, colors_by_defect, condition_ranges=None):
    sheet = workbook.create_sheet("Per Condition")
    condition_ranges = condition_ranges if condition_ranges is not None else {}
    stations = _condition_station_list(data)
    periods_by_station = _group_by(data.get("condition_periods") or [], "source_station")
    row_cursor = 1

    if not stations:
        _append_table(sheet, "tblConditionEmpty", ["Fecha", "Total NOK"], [])
        _fit_columns(sheet)
        return sheet

    for station_idx, station in enumerate(stations, start=1):
        label = _station_name(station)
        sheet.cell(row=row_cursor, column=1, value=f"{label} - Defectos dia a dia")
        sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 1

        periods = periods_by_station.get(station) or []
        classes = _condition_classes(periods)
        headers = ["Fecha", *classes, "Total NOK"] if classes else ["Fecha", "Total NOK"]
        daily_rows = _condition_daily_rows(periods, classes)
        header_row, daily_max_row, daily_max_col, total_row = _append_condition_daily_table(
            sheet,
            f"tblConditionDaily{station_idx}",
            headers,
            daily_rows,
            row_cursor,
        )
        condition_ranges[station] = _condition_daily_meta(
            sheet,
            station,
            label,
            classes,
            daily_rows,
            header_row,
            daily_max_row,
            daily_max_col,
            total_row,
        )
        _add_condition_pie_chart(
            sheet,
            label,
            classes,
            header_row,
            total_row,
            colors_by_defect,
            f"{get_column_letter(daily_max_col + 2)}{header_row}",
        )
        row_cursor = (total_row or daily_max_row) + 3

    sheet.freeze_panes = "A2"
    _fit_columns(sheet)
    return sheet


def _append_combined_conditions_section(sheet, data, colors_by_defect, condition_ranges=None):
    combined_data = _combined_as_station_data(data)
    condition_ranges = condition_ranges if condition_ranges is not None else {}
    stations = _condition_station_list(combined_data)
    if not stations:
        return

    periods_by_station = _group_by(combined_data.get("condition_periods") or [], "source_station")
    row_cursor = sheet.max_row + 3
    sheet.cell(row=row_cursor, column=1, value="Combinado LEFT+RIGHT - Per Condition")
    sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
    row_cursor += 2

    for station_idx, station in enumerate(stations, start=1):
        label = _station_name(station)
        sheet.cell(row=row_cursor, column=1, value=f"{label} - Defectos dia a dia")
        sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 1

        periods = periods_by_station.get(station) or []
        classes = _condition_classes(periods)
        headers = ["Fecha", *classes, "Total NOK"] if classes else ["Fecha", "Total NOK"]
        daily_rows = _condition_daily_rows(periods, classes)
        header_row, daily_max_row, daily_max_col, total_row = _append_condition_daily_table(
            sheet,
            f"tblCombinedConditionDaily{station_idx}",
            headers,
            daily_rows,
            row_cursor,
        )
        condition_ranges[station] = _condition_daily_meta(
            sheet,
            station,
            label,
            classes,
            daily_rows,
            header_row,
            daily_max_row,
            daily_max_col,
            total_row,
        )
        _add_condition_pie_chart(
            sheet,
            label,
            classes,
            header_row,
            total_row,
            colors_by_defect,
            f"{get_column_letter(daily_max_col + 2)}{header_row}",
        )
        row_cursor = (total_row or daily_max_row) + 3

    _fit_columns(sheet)


def _write_top3_sheet(workbook, data, colors_by_defect, condition_ranges=None):
    sheet = workbook.create_sheet("Top 3 Historico")
    condition_ranges = condition_ranges or {}
    stations = sorted(condition_ranges, key=_station_name)
    row_cursor = 1

    if not stations:
        _append_table(sheet, "tblTop3Empty", ["Top", "Class Name", "NOK acumulado"], [])
        _fit_columns(sheet)
        return sheet

    for station_idx, station in enumerate(stations, start=1):
        condition_meta = condition_ranges[station]
        label = condition_meta["label"]
        classes = _top_condition_classes(condition_meta)

        sheet.cell(row=row_cursor, column=1, value=f"{label} - Top 3 NOK por dia")
        sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 1

        summary_rows = [
            [idx + 1, name, condition_meta["totals"].get(name, 0)]
            for idx, name in enumerate(classes)
        ]
        summary_header, summary_max_row, _ = _append_table(
            sheet,
            f"tblTop3Totals{station_idx}",
            ["Top", "Class Name", "NOK acumulado"],
            summary_rows,
            start_row=row_cursor,
        )
        _add_top3_condition_chart(
            sheet,
            condition_meta,
            classes,
            colors_by_defect,
            "E" + str(summary_header),
        )
        row_cursor = max(summary_max_row + 3, summary_header + 16)

    sheet._top3_next_row = row_cursor
    sheet.freeze_panes = "A2"
    _fit_columns(sheet)
    return sheet


def _append_combined_top3_section(sheet, data, colors_by_defect, condition_ranges=None):
    condition_ranges = condition_ranges or {}
    stations = sorted(condition_ranges, key=_station_name)
    if not stations:
        return

    row_cursor = max(sheet.max_row + 3, getattr(sheet, "_top3_next_row", 0))
    sheet.cell(row=row_cursor, column=1, value="Combinado LEFT+RIGHT - Top 3 Historico")
    sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
    row_cursor += 2

    for station_idx, station in enumerate(stations, start=1):
        condition_meta = condition_ranges[station]
        label = condition_meta["label"]
        classes = _top_condition_classes(condition_meta)

        sheet.cell(row=row_cursor, column=1, value=f"{label} - Top 3 NOK por dia")
        sheet.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 1

        summary_rows = [[idx + 1, name, condition_meta["totals"].get(name, 0)] for idx, name in enumerate(classes)]
        summary_header, summary_max_row, _ = _append_table(
            sheet,
            f"tblCombinedTop3Totals{station_idx}",
            ["Top", "Class Name", "NOK acumulado"],
            summary_rows,
            start_row=row_cursor,
        )
        _add_top3_condition_chart(
            sheet,
            condition_meta,
            classes,
            colors_by_defect,
            "E" + str(summary_header),
        )
        row_cursor = max(summary_max_row + 3, summary_header + 16)

    sheet._top3_next_row = row_cursor
    _fit_columns(sheet)


def build_workbook(report_params, data):
    workbook = Workbook()
    data = data or {}
    colors_by_defect = _defect_color_map(data)
    combined_colors_by_defect = _defect_color_map(_combined_as_station_data(data))
    condition_ranges = {}
    combined_condition_ranges = {}
    daily_sheet = _write_daily_sheet(workbook, data, report_params)
    _fit_columns(daily_sheet)
    conditions_sheet = _write_conditions_sheet(workbook, data, colors_by_defect, condition_ranges)
    _append_combined_conditions_section(conditions_sheet, data, combined_colors_by_defect, combined_condition_ranges)
    top3_sheet = _write_top3_sheet(workbook, data, colors_by_defect, condition_ranges)
    _append_combined_top3_section(top3_sheet, data, combined_colors_by_defect, combined_condition_ranges)
    return workbook


def save_workbook(workbook, output_dir):
    report_dir = Path(output_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    output_path = report_dir / f"vision_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    workbook.save(output_path)
    return output_path


def generate_report(report_params, session=None):
    data = fetch_report_data(report_params, session=session)
    workbook = build_workbook(report_params, data)
    return save_workbook(workbook, report_params.output_dir)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Generate Vision Excel report based on the frontend reject summary.")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--days", type=int, default=DEFAULT_DAYS)
    parser.add_argument("--start-at")
    parser.add_argument("--end-at")
    parser.add_argument("--source-station")
    parser.add_argument("--part-number", dest="part_numbers", action="append")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv=None):
    try:
        args = parse_args(argv)
        if args.days < 1:
            raise ReportError("--days must be greater than zero")
        report_params = build_params(args)
        output_path = generate_report(report_params)
    except ReportError as exc:
        print(f"Error: {exc}")
        return 1

    print(f"Reporte generado: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
