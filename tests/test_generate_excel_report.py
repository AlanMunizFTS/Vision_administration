import tempfile
import unittest
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook

from scripts.generate_excel_report import (
    COLORS,
    DEFAULT_DAYS,
    ReportParams,
    build_workbook,
    default_period,
    fetch_report_data,
    generate_report,
)


SAMPLE_REJECT_SUMMARY = {
    "stations": [
        {
            "source_station": "station-a",
            "total_pieces": 10,
            "ok_pieces": 7,
            "nok_pieces": 3,
            "pct_ok": 0.7,
            "pct_nok": 0.3,
        }
    ],
    "daily": [
        {
            "source_station": "station-a_LEFT",
            "reject_date": "2026-06-25",
            "total_pieces": 4,
            "ok_pieces": 3,
            "nok_pieces": 1,
            "pct_ok": 0.75,
            "pct_nok": 0.25,
        },
        {
            "source_station": "station-a_LEFT",
            "reject_date": "2026-06-26",
            "total_pieces": 6,
            "ok_pieces": 4,
            "nok_pieces": 2,
            "pct_ok": 0.667,
            "pct_nok": 0.333,
        },
        {
            "source_station": "station-a_RIGHT",
            "reject_date": "2026-06-26",
            "total_pieces": 5,
            "ok_pieces": 2,
            "nok_pieces": 3,
            "pct_ok": 0.4,
            "pct_nok": 0.6,
        },
    ],
    "condition_periods": [
        {
            "source_station": "station-a",
            "reject_date": "2026-06-25",
            "class_name": "scratch",
            "nok_pieces": 1,
            "ok_pieces": 0,
            "total_pieces": 1,
        },
        {
            "source_station": "station-a",
            "reject_date": "2026-06-26",
            "class_name": "scratch",
            "nok_pieces": 1,
            "ok_pieces": 0,
            "total_pieces": 1,
        },
        {
            "source_station": "station-a",
            "reject_date": "2026-06-26",
            "class_name": "dent",
            "nok_pieces": 1,
            "ok_pieces": 0,
            "total_pieces": 1,
        },
    ],
    "condition_totals": [
        {"source_station": "station-a", "class_name": "scratch", "nok_pieces": 2},
        {"source_station": "station-a", "class_name": "dent", "nok_pieces": 1},
    ],
    "top3_history": [
        {
            "source_station": "station-a",
            "class_name": "scratch",
            "total_nok_pieces": 2,
            "class_rank": 1,
            "reject_date": "2026-06-25",
            "nok_pieces": 1,
        },
        {
            "source_station": "station-a",
            "class_name": "scratch",
            "total_nok_pieces": 2,
            "class_rank": 1,
            "reject_date": "2026-06-26",
            "nok_pieces": 1,
        },
        {
            "source_station": "station-a",
            "class_name": "dent",
            "total_nok_pieces": 1,
            "class_rank": 2,
            "reject_date": "2026-06-26",
            "nok_pieces": 1,
        },
    ],
    "combined": {
        "stations": [
            {
                "station_pair": "station-a",
                "source_stations": ["station-a_LEFT", "station-a_RIGHT"],
                "total_pieces": 8,
                "ok_pieces": 5,
                "nok_pieces": 3,
                "pct_ok": 0.625,
                "pct_nok": 0.375,
            }
        ],
        "daily": [
            {
                "station_pair": "station-a",
                "reject_date": "2026-06-26",
                "total_pieces": 8,
                "ok_pieces": 5,
                "nok_pieces": 3,
                "pct_ok": 0.625,
                "pct_nok": 0.375,
            }
        ],
        "condition_periods": [
            {
                "station_pair": "station",
                "reject_date": "2026-06-26",
                "class_name": "dent",
                "nok_pieces": 3,
                "ok_pieces": 0,
                "total_pieces": 3,
            }
        ],
        "condition_totals": [
            {"station_pair": "station", "class_name": "dent", "nok_pieces": 3},
        ],
        "top3_history": [
            {
                "station_pair": "station",
                "class_name": "dent",
                "total_nok_pieces": 3,
                "class_rank": 1,
                "reject_date": "2026-06-26",
                "nok_pieces": 3,
            }
        ],
    },
}


class FakeResponse:
    def __init__(self, payload):
        self.payload = payload

    def raise_for_status(self):
        return None

    def json(self):
        return self.payload


class FakeSession:
    def __init__(self):
        self.calls = []

    def get(self, url, params=None, timeout=None):
        self.calls.append((url, dict(params or {}), timeout))
        if url.endswith("/health"):
            return FakeResponse({"api": "ok", "database": "ok"})
        if url.endswith("/api/v1/reject-summary"):
            return FakeResponse(SAMPLE_REJECT_SUMMARY)
        raise AssertionError(f"Unexpected URL: {url}")


class GenerateExcelReportTests(unittest.TestCase):
    def make_params(self, output_dir="reports", source_station=None, part_numbers=None):
        return ReportParams(
            api_url="http://testserver",
            start_at="2026-06-19 00:00:00",
            end_at="2026-06-26 23:59:59",
            source_station=source_station,
            part_numbers=part_numbers,
            output_dir=output_dir,
        )

    def test_default_period_uses_last_seven_days_full_day_bounds(self):
        start_at, end_at = default_period()

        self.assertEqual(DEFAULT_DAYS, 7)
        self.assertEqual((end_at.date() - start_at.date()).days, 7)
        self.assertEqual((start_at.hour, start_at.minute, start_at.second), (0, 0, 0))
        self.assertEqual((end_at.hour, end_at.minute, end_at.second), (23, 59, 59))

    def test_fetch_report_data_uses_frontend_reject_summary_endpoint(self):
        session = FakeSession()
        data = fetch_report_data(self.make_params(source_station="station-a"), session=session)

        self.assertEqual(data["daily"][0]["reject_date"], "2026-06-25")
        self.assertEqual(
            session.calls,
            [
                ("http://testserver/health", {}, 20),
                (
                    "http://testserver/api/v1/reject-summary",
                    {
                        "start_at": "2026-06-19 00:00:00",
                        "end_at": "2026-06-26 23:59:59",
                        "source_station": "station-a",
                    },
                    20,
                ),
            ],
        )

    def test_build_workbook_creates_frontend_sheets_tables_and_charts(self):
        workbook = build_workbook(self.make_params(), SAMPLE_REJECT_SUMMARY)
        daily = workbook["Por dia"]

        self.assertEqual(workbook.sheetnames, ["Por dia", "Per Condition", "Top 3 Historico"])
        self.assertEqual(daily.tables, {})
        self.assertTrue(workbook["Per Condition"].tables)
        self.assertTrue(workbook["Top 3 Historico"].tables)
        self.assertEqual(daily.freeze_panes, "A8")
        self.assertEqual(daily["A1"].value, "Filtros aplicados")
        self.assertEqual(daily["A2"].value, "Inicio: 2026-06-19 00:00:00")
        self.assertEqual(daily["C2"].value, "Fin: 2026-06-26 23:59:59")
        self.assertEqual(daily["A3"].value, "Estacion: Todos")
        self.assertEqual(daily["C3"].value, "Part Number: Todos")
        self.assertIn("B5:P5", [str(item) for item in daily.merged_cells.ranges])
        self.assertIn("B6:F6", [str(item) for item in daily.merged_cells.ranges])
        self.assertIn("G6:K6", [str(item) for item in daily.merged_cells.ranges])
        self.assertIn("L6:P6", [str(item) for item in daily.merged_cells.ranges])
        self.assertEqual(daily["B5"].value, "station-a")
        self.assertEqual(daily["B6"].value, "Left")
        self.assertEqual(daily["G6"].value, "Right")
        self.assertEqual(daily["L6"].value, "Combinado")
        self.assertEqual([daily.cell(row=7, column=col).value for col in range(1, 7)], ["Date", "OK", "NOK", "Total", "% OK", "% NOK"])
        self.assertEqual(daily["A8"].value, "2026-06-25")
        self.assertEqual(daily["B8"].value, 3)
        self.assertEqual(daily["G8"].value, "")
        self.assertEqual(daily["L9"].value, 5)
        self.assertEqual(daily["P9"].value, 0.375)
        self.assertEqual(daily["F9"].number_format, "0.00%")
        self.assertEqual(daily["K9"].number_format, "0.00%")
        self.assertEqual(daily["P9"].number_format, "0.00%")
        self.assertEqual(daily["F10"].value, "=AVERAGE(F8:F9)")
        self.assertEqual(daily["K10"].value, "=AVERAGE(K8:K9)")
        self.assertEqual(daily["P10"].value, "=AVERAGE(P8:P9)")
        self.assertGreaterEqual(len(daily.conditional_formatting), 4)
        self.assertEqual(len(daily._charts), 1)
        self.assertGreaterEqual(len(workbook["Per Condition"]._charts), 1)
        self.assertGreaterEqual(len(workbook["Top 3 Historico"]._charts), 1)
        self.assertEqual(daily._charts[0].y_axis.scaling.min, 0)
        self.assertEqual(daily._charts[0].y_axis.scaling.max, 1)

    def test_build_workbook_displays_part_number_filters(self):
        workbook = build_workbook(self.make_params(part_numbers=["PN-1", "PN-2"]), SAMPLE_REJECT_SUMMARY)

        daily = workbook["Por dia"]
        self.assertEqual(daily["C3"].value, "Part Number: PN-1, PN-2")
        self.assertEqual(daily["A8"].value, "2026-06-25")

    def test_build_workbook_displays_art_endform_stations_as_tesla_names(self):
        data = {
            "stations": [],
            "daily": [
                {
                    "source_station": "ART_ENDFORM_1859_LEFT",
                    "reject_date": "2026-06-25",
                    "total_pieces": 4,
                    "ok_pieces": 3,
                    "nok_pieces": 1,
                    "pct_ok": 0.75,
                    "pct_nok": 0.25,
                },
                {
                    "source_station": "ART_ENDFORM_1859_RIGHT",
                    "reject_date": "2026-06-25",
                    "total_pieces": 6,
                    "ok_pieces": 4,
                    "nok_pieces": 2,
                    "pct_ok": 0.667,
                    "pct_nok": 0.333,
                },
            ],
            "condition_periods": [
                {
                    "source_station": "ART_ENDFORM_1859_LEFT",
                    "reject_date": "2026-06-25",
                    "class_name": "scratch",
                    "nok_pieces": 1,
                    "ok_pieces": 0,
                    "total_pieces": 1,
                }
            ],
            "condition_totals": [
                {"source_station": "ART_ENDFORM_1859_LEFT", "class_name": "scratch", "nok_pieces": 1},
            ],
            "top3_history": [
                {
                    "source_station": "ART_ENDFORM_1859_LEFT",
                    "class_name": "scratch",
                    "total_nok_pieces": 1,
                    "class_rank": 1,
                    "reject_date": "2026-06-25",
                    "nok_pieces": 1,
                }
            ],
            "combined": {
                "stations": [],
                "daily": [
                    {
                        "station_pair": "ART_ENDFORM_1859",
                        "reject_date": "2026-06-25",
                        "total_pieces": 10,
                        "ok_pieces": 7,
                        "nok_pieces": 3,
                        "pct_ok": 0.7,
                        "pct_nok": 0.3,
                    }
                ],
                "condition_periods": [
                    {
                        "station_pair": "ART_ENDFORM_1859",
                        "reject_date": "2026-06-25",
                        "class_name": "scratch",
                        "nok_pieces": 3,
                        "ok_pieces": 0,
                        "total_pieces": 3,
                    }
                ],
                "condition_totals": [
                    {"station_pair": "ART_ENDFORM_1859", "class_name": "scratch", "nok_pieces": 3},
                ],
                "top3_history": [
                    {
                        "station_pair": "ART_ENDFORM_1859",
                        "class_name": "scratch",
                        "total_nok_pieces": 3,
                        "class_rank": 1,
                        "reject_date": "2026-06-25",
                        "nok_pieces": 3,
                    }
                ],
            },
        }

        workbook = build_workbook(self.make_params(), data)
        daily = workbook["Por dia"]
        conditions = workbook["Per Condition"]
        top3 = workbook["Top 3 Historico"]

        self.assertEqual(daily["B5"].value, "Tesla 1")
        self.assertEqual(daily["B6"].value, "Left")
        self.assertEqual(daily["G6"].value, "Right")
        self.assertEqual(daily["L6"].value, "Combinado")
        self.assertEqual(conditions["A1"].value, "Tesla 1 - Left - Defectos dia a dia")
        self.assertIn(
            "Tesla 1 - Defectos dia a dia",
            [cell.value for row in conditions.iter_rows() for cell in row],
        )
        self.assertEqual(top3["A1"].value, "Tesla 1 - Left - Top 3 NOK por dia")
        self.assertIn(
            "Tesla 1 - Top 3 NOK por dia",
            [cell.value for row in top3.iter_rows() for cell in row],
        )

    def test_build_workbook_uppercases_and_alpha_sorts_defect_names(self):
        workbook = build_workbook(self.make_params(), SAMPLE_REJECT_SUMMARY)

        conditions = workbook["Per Condition"]
        top3 = workbook["Top 3 Historico"]
        self.assertEqual(conditions["B2"].value, "DENT")
        self.assertEqual(conditions["C2"].value, "SCRATCH")
        self.assertEqual(conditions["A5"].value, "Total")
        self.assertEqual(conditions["B5"].value, "=SUM(B3:B4)")
        self.assertEqual(conditions["C5"].value, "=SUM(C3:C4)")
        self.assertEqual(conditions._charts[0].series[0].val.numRef.f, "'Per Condition'!$B$5:$C$5")
        self.assertEqual(conditions._charts[0].series[0].cat.numRef.f, "'Per Condition'!$B$2:$C$2")
        self.assertEqual(top3["B3"].value, "SCRATCH")
        self.assertEqual(top3["B4"].value, "DENT")
        self.assertNotIn("tblTop3History1", top3.tables)
        self.assertEqual(top3._charts[0].series[0].val.numRef.f, "'Per Condition'!$C$3:$C$4")
        self.assertEqual(top3._charts[0].series[1].val.numRef.f, "'Per Condition'!$B$3:$B$4")
        self.assertEqual(top3._charts[0].series[0].cat.numRef.f, "'Per Condition'!$A$3:$A$4")
        self.assertEqual(top3._charts[0].x_axis.tickLblPos, "low")
        self.assertEqual(top3._charts[0].y_axis.tickLblPos, "nextTo")
        self.assertEqual(top3._charts[0].y_axis.numFmt.formatCode, "0")
        self.assertEqual(top3._charts[0].y_axis.scaling.min, 0)
        self.assertEqual(top3._charts[0].y_axis.scaling.max, 1)
        condition_colors = [
            point.spPr.solidFill.srgbClr
            for point in conditions._charts[0].series[0].data_points
        ]
        top3_colors = [
            series.graphicalProperties.solidFill.srgbClr
            for series in top3._charts[0].series
        ]
        self.assertEqual(top3_colors, [condition_colors[1], condition_colors[0]])

    def test_combined_sections_use_frontend_independent_color_map(self):
        data = deepcopy(SAMPLE_REJECT_SUMMARY)
        data["combined"]["condition_totals"] = [
            {"station_pair": "station", "class_name": "scratch", "nok_pieces": 3},
        ]
        data["combined"]["condition_periods"] = [
            {
                "station_pair": "station",
                "reject_date": "2026-06-26",
                "class_name": "scratch",
                "nok_pieces": 3,
                "ok_pieces": 0,
                "total_pieces": 3,
            }
        ]
        data["combined"]["top3_history"] = [
            {
                "station_pair": "station",
                "class_name": "scratch",
                "total_nok_pieces": 3,
                "class_rank": 1,
                "reject_date": "2026-06-26",
                "nok_pieces": 3,
            }
        ]

        workbook = build_workbook(self.make_params(), data)

        combined_condition_chart = workbook["Per Condition"]._charts[1]
        combined_top3_chart = workbook["Top 3 Historico"]._charts[1]
        self.assertEqual(combined_condition_chart.series[0].data_points[0].spPr.solidFill.srgbClr, COLORS[0])
        self.assertEqual(combined_top3_chart.series[0].graphicalProperties.solidFill.srgbClr, COLORS[0])

    def test_generate_report_writes_xlsx_file(self):
        session = FakeSession()
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = generate_report(self.make_params(output_dir=tmp_dir), session=session)

            self.assertTrue(Path(output_path).exists())
            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["Por dia", "Per Condition", "Top 3 Historico"])
            self.assertEqual(workbook["Por dia"]["A8"].value, "2026-06-25")

    def test_empty_data_still_creates_workbook(self):
        params = self.make_params()
        data = {
            "stations": [],
            "daily": [],
            "condition_periods": [],
            "condition_totals": [],
            "top3_history": [],
        }

        workbook = build_workbook(params, data)

        self.assertEqual(workbook.sheetnames, ["Por dia", "Per Condition", "Top 3 Historico"])
        self.assertEqual(workbook["Por dia"]["A8"].value, "Sin datos")
        self.assertEqual(workbook["Por dia"].tables, {})
        self.assertEqual(workbook["Por dia"].freeze_panes, "A8")
        self.assertTrue(workbook["Per Condition"].tables)
        self.assertTrue(workbook["Top 3 Historico"].tables)


if __name__ == "__main__":
    unittest.main()
